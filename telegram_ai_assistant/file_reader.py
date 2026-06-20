"""
Hujjat fayllari mazmunini matnqa aylantiruvchi modul.
Qo'llab-quvvatlanadigan formatlar: PDF, DOCX, XLSX/XLS, TXT, CSV.
"""

import logging
from pathlib import Path
from typing import Optional

log = logging.getLogger("telegram_ai_assistant.file_reader")

MAX_CHARS = 15_000  # Claude ga yuboriladigan maksimal belgi soni

SUPPORTED_EXT = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".txt", ".csv", ".md"}


def is_supported(filename: str) -> bool:
    return Path(filename).suffix.lower() in SUPPORTED_EXT


def _read_pdf(path: Path) -> str:
    import pdfplumber
    pages = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if text.strip():
                pages.append(f"[Bet {i+1}]\n{text.strip()}")
            if sum(len(p) for p in pages) > MAX_CHARS:
                pages.append(f"... (jami {len(pdf.pages)} bet, qolganlar o'tkazildi)")
                break
    return "\n\n".join(pages)


def _read_docx(path: Path) -> str:
    from docx import Document
    doc = Document(path)
    paragraphs = []
    total = 0
    for para in doc.paragraphs:
        t = para.text.strip()
        if not t:
            continue
        paragraphs.append(t)
        total += len(t)
        if total > MAX_CHARS:
            paragraphs.append("... (qolgan qismlar o'tkazildi)")
            break
    return "\n".join(paragraphs)


def _read_xlsx(path: Path) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    parts = []
    total = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append(" | ".join(cells))
        if rows:
            block = f"[Varaq: {sheet_name}]\n" + "\n".join(rows)
            parts.append(block)
            total += len(block)
        if total > MAX_CHARS:
            parts.append("... (qolgan varaqlar o'tkazildi)")
            break
    return "\n\n".join(parts)


def _read_text(path: Path) -> str:
    for enc in ("utf-8", "utf-8-sig", "cp1251", "latin-1"):
        try:
            text = path.read_text(encoding=enc)
            return text[:MAX_CHARS] + ("..." if len(text) > MAX_CHARS else "")
        except (UnicodeDecodeError, LookupError):
            continue
    return "[Fayl kodlashini aniqlab bo'lmadi]"


def extract_text(path: Path) -> Optional[str]:
    """Fayl mazmunini matnqa aylantiradi. Xato bo'lsa None qaytaradi."""
    ext = path.suffix.lower()
    try:
        if ext == ".pdf":
            text = _read_pdf(path)
        elif ext in (".docx", ".doc"):
            text = _read_docx(path)
        elif ext in (".xlsx", ".xls"):
            text = _read_xlsx(path)
        elif ext in (".txt", ".csv", ".md"):
            text = _read_text(path)
        else:
            return None
        return text.strip() or "(fayl bo'sh)"
    except Exception as exc:
        log.warning("Fayl o'qish xatosi %s: %s", path.name, exc)
        return f"[Fayl o'qishda xatolik: {exc}]"


def label_for_ext(ext: str) -> str:
    return {
        ".pdf":  "📄 PDF",
        ".docx": "📝 Word",
        ".doc":  "📝 Word",
        ".xlsx": "📊 Excel",
        ".xls":  "📊 Excel",
        ".txt":  "📃 Matn",
        ".csv":  "📊 CSV",
        ".md":   "📃 Markdown",
    }.get(ext.lower(), "📎 Fayl")
