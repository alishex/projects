from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.config import settings
from app.database import db, now_str

HEADERS = [
    "ID", "F.I.SH.", "Telegram ID", "Username", "Telefon", "Tug'ilgan sana", "Hudud", "Tuman", "Manzil",
    "Tanlangan lavozim", "Vakansiya ID", "Reglament versiya snapshot", "Qabul qilingan bo'lim", "Ta'lim", "Talabami", "Til darajalari", "Kompyuter darajasi",
    "Ish tajribasi bor/yo'q", "Ish joylari soni", "Oldingi lavozimlar", "Ketish sabablari", "Tasdiqlovchi kontaktlar",
    "AI interview score", "AI interview grade", "AI summary", "Status", "Intervyu sanasi", "Intervyu manzili",
    "Stajirovka statusi", "Tugallangan darslar soni", "Har bir dars test o'rtacha bali", "Yakuniy 30 test natijasi",
    "Yakuniy AI tavsiya", "Clockster ID", "Clockster ism", "Oxirgi kelgan vaqt", "Oxirgi ketgan vaqt", "Clockster oxirgi sync", "Admin izohi", "Ariza sanasi"
]


def export_candidates_excel() -> Path:
    settings.export_dir.mkdir(parents=True, exist_ok=True)
    path = settings.export_dir / f"allmax_hr_export_{now_str().replace(':','-').replace(' ','_')}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    candidates = []
    for sec in ["new", "ai_rejected", "medium", "excellent", "invited", "accepted", "rejected", "reserve"]:
        candidates.extend(db.list_candidates(sec, limit=10000))
    seen = set()
    for c in candidates:
        if c["id"] in seen:
            continue
        seen.add(c["id"])
        work = db.candidate_work_experiences(c["id"])
        employee = db.get_employee_by_candidate(c["id"])
        tests = db.lesson_test_stats(employee["id"]) if employee else []
        avg = round(sum(float(t.get("percentage") or 0) for t in tests) / len(tests), 2) if tests else ""
        final_result = ""
        final_ai = ""
        if employee:
            finals = [f for f in db.final_tests(limit=10000) if f.get("employee_id") == employee["id"]]
            if finals:
                f = finals[0]
                final_result = f"{f.get('correct_answers')}/{f.get('total_questions')} — {f.get('percentage')}%"
                final_ai = f.get("admin_recommendation") or f.get("ai_summary") or ""
        interviews = ""
        location = ""
        # lightweight lookup by current schema not direct method list; skip if absent
        latest_att = db.clockster_attendance_for_employee(employee["id"], limit=1)[0] if employee and db.clockster_attendance_for_employee(employee["id"], limit=1) else {}
        row = [
            c.get("id"), c.get("full_name"), c.get("telegram_id"), c.get("username"), c.get("phone"), c.get("birth_date"),
            c.get("region"), c.get("district"), c.get("address"), c.get("position"), c.get("vacancy_id"), c.get("regulation_version_snapshot"), c.get("accepted_department"), c.get("education_level"),
            c.get("is_student"), f"UZ:{c.get('uzbek_level')} RU:{c.get('russian_level')} EN:{c.get('english_level')}", c.get("computer_level"),
            c.get("has_experience"), len(work), "; ".join(str(w.get("position") or "") for w in work),
            "; ".join(str(w.get("leaving_reason") or "") for w in work),
            "; ".join(f"{w.get('reference_name') or ''} {w.get('reference_phone') or ''}" for w in work),
            c.get("ai_score"), c.get("ai_grade"), c.get("ai_summary"), c.get("status"), interviews, location,
            employee.get("status") if employee else "", db.completed_lessons_count(employee["id"]) if employee else "", avg,
            final_result, final_ai,
            employee.get("clockster_employee_id") if employee else "", employee.get("clockster_employee_name") if employee else "",
            latest_att.get("check_in", ""), latest_att.get("check_out", ""), employee.get("clockster_last_sync_at") if employee else "",
            c.get("admin_note"), c.get("created_at")
        ]
        ws.append(row)
    for col in range(1, len(HEADERS) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = 18 if col not in [23, 31, 32] else 35
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
    return path
